import boto3
from boto3.dynamodb.conditions import Key
from datetime import datetime, time
import re
from openpyxl import Workbook, load_workbook
import time as time_module
import os
from openpyxl.styles import PatternFill
import smtplib
from email.mime.text import MIMEText

SENDER_EMAIL = "aman@ai4mtech.com"
PASSWORD = "mgreprmdmdfdlaqc"

def validate_vehicle_number(vehicle_number):
    valid_state_codes = [
        'AN', 'AP', 'AR', 'AS', 'BR', 'CH', 'CT', 'DL', 'GA', 'GJ', 'HR', 'HP', 'JK', 'JH', 'KA', 
        'KL', 'MP', 'MH', 'MN', 'ML', 'MZ', 'NL', 'OR', 'PY', 'PB', 'RJ', 'SK', 'TN', 'TS', 'TR', 
        'UP', 'UT', 'WB', 'DD','UK'
    ]
    vehicle_number = vehicle_number.replace(" ", "").upper()
    if not re.match(r"^[A-Z]{2}\d{2}[A-Z]{1,2}\d{4}$", vehicle_number):
        return False
    state_code = vehicle_number[:2]
    if state_code not in valid_state_codes:
        return False
    return True

def send_email_alert(vehicle_number, timestamp):
    recipient_email = "yogita@ai4mtech.com"
    subject = "Incorrect Vehicle Number Detected"
    body = f"Vehicle Number: {vehicle_number}\nTimestamp: {timestamp}\nStatus: incorrect"
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = SENDER_EMAIL
    msg['To'] = recipient_email

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(SENDER_EMAIL, PASSWORD)
            server.send_message(msg)
        print(f"Email alert sent to {recipient_email} for vehicle {vehicle_number}")
    except Exception as e:
        print(f"Failed to send email: {str(e)}")

def fetch_and_write():
    dynamodb = boto3.resource('dynamodb')
    table = dynamodb.Table('rhenus_vehicle_number')

    current_date = datetime.now().date()
    filename = f"vehicle_entry_status_{current_date.strftime('%Y-%m-%d')}.xlsx"

    # Open or create Excel file for current date
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Entry Status"
        ws.append(["vehicle_number", "entry_image_url", "status", "timestamp", "whatsapp"])

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    while True:
        try:
            now = datetime.now()
            today = now.date()

            # If date changed, save current file and start a new one
            if today != current_date:
                wb.save(filename)

                current_date = today
                filename = f"vehicle_entry_status_{current_date.strftime('%Y-%m-%d')}.xlsx"

                if os.path.exists(filename):
                    wb = load_workbook(filename)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Entry Status"
                    ws.append(["vehicle_number", "entry_image_url", "status", "timestamp", "whatsapp"])

            start_date = datetime.combine(today, time.min)
            end_date = datetime.combine(today, time.max)

            response = table.query(
                KeyConditionExpression=Key('vechicle_number_type').eq('gate_in') & 
                                       Key('timestamp').between(start_date.isoformat(), end_date.isoformat()),
                FilterExpression=Key('vehicle_type').eq('Truck'),
                ScanIndexForward=False
            )
            items = response.get('Items', [])

            # Track existing entries to avoid duplicates
            existing_entries = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                vnum = str(row[0]).strip() if row[0] is not None else ""
                ts = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                existing_entries.add((vnum, ts))

            new_rows = 0
            for item in items:
                vehicle_number = str(item.get("vehicle_number", "")).strip()
                entry_image_url = item.get("entry_image_url", "")
                timestamp = str(item.get("timestamp", "")).strip()
                status = "correct" if validate_vehicle_number(vehicle_number) else "incorrect"
                key = (vehicle_number, timestamp)
                if key in existing_entries:
                    continue

                ws.append([vehicle_number, entry_image_url, status, timestamp, "false"])
                if status == "incorrect":
                    for cell in ws[ws.max_row]:
                        cell.fill = yellow_fill
                new_rows += 1

            # Send alerts for new incorrect entries and update Whatsapp flag
            for row in ws.iter_rows(min_row=2):
                vehicle_number = row[0].value
                timestamp = row[3].value
                status = row[2].value
                whatsapp = row[4].value
                if status == "incorrect" and whatsapp == "false":
                    send_email_alert(vehicle_number, timestamp)
                    row[4].value = "true"

            if new_rows > 0:
                # Sort entries by timestamp descending
                data = list(ws.iter_rows(min_row=2, values_only=True))
                data.sort(key=lambda x: x[3] if x[3] else "", reverse=True)
                ws.delete_rows(2, ws.max_row)
                for row_data in data:
                    ws.append(row_data)
                    if row_data[2] == "incorrect":
                        for cell in ws[ws.max_row]:
                            cell.fill = yellow_fill
                wb.save(filename)
                print(f"Added {new_rows} new entries to {filename}")
            else:
                # Save even if no new rows to update Whatsapp flags or minor changes
                wb.save(filename)

            time_module.sleep(300)  # Sleep 5 minutes before next fetch

        except Exception as e:
            print(f"Error: {e}")
            time_module.sleep(60)  # Sleep 1 minute before retry

if __name__ == "__main__":
    fetch_and_write()
