import os
import sys
import subprocess
import socket
import smtplib
import json
import psutil
import shutil
import platform
import importlib
import pkg_resources
import re
import time
import boto3
import decimal
from email.message import EmailMessage
from datetime import datetime
from typing import Dict, Any, List, Tuple
from threading import Thread, Event
from decimal import Decimal

# Check if we're on Windows
IS_WINDOWS = platform.system() == "Windows"
IS_LINUX = platform.system().lower() == "linux"

# Add boto3 to required packages
REQUIRED_PACKAGES = {
    'openpyxl': 'openpyxl',
    'boto3': 'boto3',
}

# Initialize JTOP_AVAILABLE early
JTOP_AVAILABLE = False
IS_JETSON = False
try:
    if IS_LINUX:
        with open("/proc/device-tree/model", "r") as f:
            if "NVIDIA Jetson" in f.read():
                IS_JETSON = True
except Exception:
    pass

# Jetson-specific package check
if not IS_WINDOWS and platform.machine() in ['aarch64', 'armv7l']:
    try:
        with open('/proc/device-tree/model', 'r') as f:
            model = f.read().lower()
            if 'jetson' in model:
                try:
                    import jtop
                    JTOP_AVAILABLE = True
                    print("jtop is already available")
                except ImportError:
                    REQUIRED_PACKAGES['jtop'] = 'jetson-stats'
                    JTOP_AVAILABLE = False
    except:
        pass

def install_package(package_name, import_name=None):
    """Install a package using pip only if not available"""
    if import_name is None:
        import_name = package_name
        
    try:
        importlib.import_module(import_name)
        print(f"{package_name} is already installed.")
        return True
    except ImportError:
        print(f"Installing {package_name}...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
            return True
        except Exception as e:
            print(f"Note: {package_name} may already be installed or unavailable: {e}")
            return False

# Install missing packages
for pip_name, import_name in REQUIRED_PACKAGES.items():
    install_package(pip_name, import_name)

# Now import the required packages
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import boto3
    from botocore.exceptions import ClientError, NoCredentialsError
    
    # Try to import jtop if we're on a Jetson device and it was added to required packages
    if 'jtop' in REQUIRED_PACKAGES:
        try:
            import jtop
            JTOP_AVAILABLE = True
        except ImportError:
            JTOP_AVAILABLE = False
            print("jtop not available, skipping Jetson-specific metrics")
except ImportError as e:
    print(f"Failed to import required packages: {e}")
    sys.exit(1)

# ───────────────────────────────────────────────
# Load Config
try:
    with open("config_monitor.json", "r") as f:
        CONFIG = json.load(f)
except FileNotFoundError:
    # Create a default config if none exists
    CONFIG = {
        "AIC_ID": "UNKNOWN",
        "AIC_NAME": "UNKNOWN",
        "PROJECT": "UNKNOWN",
        "SYSTEM": {
            "IP": "8.8.8.8",
            "NAME": "Google DNS"
        },
        "PEER_IPS": {},
        "CAMERA_IPS": {},
        "EMAIL": {
            "SENDER": "your-email@gmail.com",
            "PASSWORD": "your-app-password",
            "RECEIVERS": ["receiver@example.com"]
        },
        "AWS": {
            "REGION": "us-east-1",
            "ACCESS_KEY_ID": "your-access-key",
            "SECRET_ACCESS_KEY": "your-secret-key"
        },
        "DYNAMODB_TABLES": {
            "PROJECTS": "ai4m_projects",
            "CAMERAS": "ai4m_cameras",
            "SERVICES": "ai4m_services"
        },
        "CUSTOMER_NAME": "Pravin Masale",
        "LOCATION": "Hadapsar",
        "DESCRIPTION": "Warehouse Monitoring",
        "SERVICES": [],
        "PYTHON_FILES": [],
        "SEARCH_PATHS": ["/", "/home", "/var", "/opt", "/usr", "/root", "/tmp"] if not IS_WINDOWS else ["C:\\"]
    }
    with open("suhana_bhuleshwar_aic_monitor_config1.json", "w") as f:
        json.dump(CONFIG, f, indent=4)
    print("Default config.json created. Please edit it with your settings.")
    sys.exit(1)

AIC_ID = CONFIG.get("AIC_ID", "UNKNOWN")
AIC_NAME = CONFIG.get("AIC_NAME", "UNKNOWN")
PROJECT = CONFIG.get("PROJECT", "UNKNOWN")

SYSTEM_IP = CONFIG.get("SYSTEM", {}).get("IP", "8.8.8.8")
SYSTEM_NAME = CONFIG.get("SYSTEM", {}).get("NAME", "Google DNS")

IP_LIST = CONFIG.get("PEER_IPS", {})
CAMERA_IPS = CONFIG.get("CAMERA_IPS", {})

SENDER_EMAIL = CONFIG.get("EMAIL", {}).get("SENDER", "")
PASSWORD = CONFIG.get("EMAIL", {}).get("PASSWORD", "")
RECEIVER_EMAILS = CONFIG.get("EMAIL", {}).get("RECEIVERS", [])

AWS_CONFIG = CONFIG.get("AWS", {})
AWS_REGION = AWS_CONFIG.get("REGION", "ap-south-1")
AWS_ACCESS_KEY_ID = AWS_CONFIG.get("ACCESS_KEY_ID", "")
AWS_SECRET_ACCESS_KEY = AWS_CONFIG.get("SECRET_ACCESS_KEY", "")

DYNAMODB_TABLES = CONFIG.get("DYNAMODB_TABLES", {
    "PROJECTS": "ai4m_projects",
    "CAMERAS": "ai4m_cameras",
    "SERVICES": "ai4m_services"
})

CUSTOMER_NAME = CONFIG.get("CUSTOMER_NAME", "Pravin Masale")
LOCATION = CONFIG.get("LOCATION", "Hadapsar")
DESCRIPTION = CONFIG.get("DESCRIPTION", "Warehouse Monitoring")

SERVICES = CONFIG.get("SERVICES", [])
PYTHON_FILES = CONFIG.get("PYTHON_FILES", [])
SEARCH_PATHS = CONFIG.get("SEARCH_PATHS", ["/", "/home", "/var", "/opt", "/usr", "/root", "/tmp"] if not IS_WINDOWS else ["C:\\"])

ERROR_MESSAGE = "System is OFF"
RETRY_COUNT = 1
AIC_STATUS = ""
# ───────────────────────────────────────────────

# Initialize DynamoDB client
try:
    if AWS_ACCESS_KEY_ID and AWS_SECRET_ACCESS_KEY:
        dynamodb = boto3.resource(
            'dynamodb',
            region_name=AWS_REGION,
            aws_access_key_id=AWS_ACCESS_KEY_ID,
            aws_secret_access_key=AWS_SECRET_ACCESS_KEY
        )
    else:
        # Try to use default credentials (e.g., from AWS CLI or IAM role)
        dynamodb = boto3.resource('dynamodb', region_name=AWS_REGION)
    
    # Get table references
    projects_table = dynamodb.Table(DYNAMODB_TABLES["PROJECTS"])
    cameras_table = dynamodb.Table(DYNAMODB_TABLES["CAMERAS"])
    services_table = dynamodb.Table(DYNAMODB_TABLES["SERVICES"])
    
    print("DynamoDB client initialized successfully")
except (NoCredentialsError, ClientError) as e:
    print(f"Failed to initialize DynamoDB client: {e}")
    dynamodb = None
    projects_table = None
    cameras_table = None
    services_table = None

# Define styles for Excel formatting
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
NORMAL_FONT = Font(size=10)
HIGHLIGHT_FONT = Font(bold=True, color="FF0000")
SUCCESS_FONT = Font(bold=True, color="006100")

HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
SECTION_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
ONLINE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for online status

THIN_BORDER = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')

# Global flag to control the background thread
stop_event = Event()

# Global dictionary to track process start times
process_start_times = {}

def get_ip_address() -> str:
    try:
        if IS_WINDOWS:
            hostname = socket.gethostname()
            return socket.gethostbyname(hostname)
        else:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            try:
                s.connect(('10.254.254.254', 1))
                ip_address = s.getsockname()[0]
            except Exception:
                ip_address = 'N/A'
            finally:
                s.close()
            return ip_address
    except Exception:
        return 'N/A'

def calculate_percentage(part, total):
    return (part / total) * 100 if total else 0.0

def get_systemd_services() -> List[Dict[str, Any]]:
    if IS_WINDOWS:
        try:
            result = subprocess.run(
                ['sc', 'query', 'type=', 'service', 'state=', 'all'],
                stdout=subprocess.PIPE, universal_newlines=True, shell=True
            )
            services = []
            current_service = {}
            for line in result.stdout.splitlines():
                if line.strip().startswith('SERVICE_NAME:'):
                    if current_service:
                        services.append(current_service)
                    current_service = {'service_name': line.split(':', 1)[1].strip()}
                elif line.strip().startswith('STATE'):
                    parts = line.split()
                    if len(parts) >= 4:
                        current_service['active_state'] = parts[3]
            if current_service:
                services.append(current_service)
            return services
        except Exception:
            return []
    else:
        try:
            result = subprocess.run(
                ['systemctl', 'list-units', '--type=service', '--all', '--no-pager', '--no-legend'],
                stdout=subprocess.PIPE, universal_newlines=True
            )
            services = result.stdout.splitlines()
            service_list = []
            for line in services:
                parts = line.split()
                if len(parts) >= 4:
                    unit = parts[0]
                    load_state = parts[1]
                    active_state = parts[2]
                    sub_state = parts[3]
                    service_name = unit.replace('.service', '')
                    service_list.append({
                        'service_name': service_name,
                        'load_state': load_state,
                        'active_state': active_state,
                        'sub_state': sub_state
                    })
            return service_list
        except Exception:
            return []

def check_ip_status(ip_map: Dict[str, str]) -> Dict[str, Dict[str, str]]:
    ip_status = {}
    if not isinstance(ip_map, dict):
        return ip_status
    
    for ip, name in ip_map.items():
        status = "offline"
        for _ in range(RETRY_COUNT):
            try:
                if IS_WINDOWS:
                    ping_cmd = ["ping", "-n", "1", ip]
                else:
                    ping_cmd = ["ping", "-c", "1", ip]
                
                with open(os.devnull, 'w') as devnull:
                    ping_result = subprocess.run(
                        ping_cmd, 
                        stdout=devnull, 
                        stderr=devnull
                    ).returncode
                
                if ping_result == 0:
                    status = "running"
                    break
            except Exception:
                status = "stopped"
        ip_status[ip] = {"name": name, "status": status}
    return ip_status

def check_services_by_name(names: List[str]) -> Dict[str, Dict[str, Any]]:
    status = {}
    for name in names:
        try:
            if IS_WINDOWS:
                result = subprocess.run(
                    ['tasklist', '/FI', f'IMAGENAME eq {name}'], 
                    stdout=subprocess.PIPE, 
                    universal_newlines=True,
                    shell=True
                )
                is_running = name.lower() in result.stdout.lower()
                status[name] = {
                    "status": "running" if is_running else "stopped",
                    "uptime": get_process_uptime(name) if is_running else "N/A"
                }
            else:
                # Try to get PID using pgrep
                result = subprocess.run(
                    ["pgrep", "-f", name], 
                    stdout=subprocess.PIPE, 
                    stderr=subprocess.PIPE,
                    universal_newlines=True
                )
                
                if result.returncode == 0:
                    pids = result.stdout.strip().split('\n')
                    if pids and pids[0]:
                        status[name] = {
                            "status": "running",
                            "uptime": get_process_uptime(pids[0])
                        }
                    else:
                        status[name] = {"status": "stopped", "uptime": "N/A"}
                else:
                    status[name] = {"status": "stopped", "uptime": "N/A"}
        except Exception:
            status[name] = {"status": "stopped", "uptime": "N/A"}
    return status

def get_process_uptime(pid_or_name):
    """Get process uptime in a human-readable format"""
    try:
        if IS_WINDOWS:
            # For Windows, we need to use a different approach
            # This is a simplified version - may need adjustment
            return "N/A"  # Windows uptime detection needs more work
        else:
            # For Linux, use ps command to get process start time
            result = subprocess.run(
                ["ps", "-o", "etimes=", "-p", str(pid_or_name)],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                universal_newlines=True
            )
            
            if result.returncode == 0 and result.stdout.strip():
                seconds = int(result.stdout.strip())
                return format_uptime(seconds)
            
            # Alternative approach for process name instead of PID
            result = subprocess.run(
                ["ps", "-o", "etimes=", "-C", str(pid_or_name)],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                universal_newlines=True
            )
            
            if result.returncode == 0 and result.stdout.strip():
                # Get the first line (oldest process)
                lines = result.stdout.strip().split('\n')
                if lines and lines[0]:
                    seconds = int(lines[0].strip())
                    return format_uptime(seconds)
            
            return "N/A"
    except Exception:
        return "N/A"

def format_uptime(seconds):
    """Format seconds into a human-readable uptime string"""
    if seconds < 60:
        return f"{seconds}s"
    elif seconds < 3600:
        minutes = seconds // 60
        return f"{minutes}m"
    elif seconds < 86400:
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        return f"{hours}h {minutes}m"
    else:
        days = seconds // 86400
        hours = (seconds % 86400) // 3600
        return f"{days}d {hours}h"

def read_meminfo() -> Dict[str, int]:
    mem = {}
    try:
        if IS_WINDOWS:
            result = subprocess.run(
                ['wmic', 'OS', 'get', 'TotalVisibleMemorySize,FreePhysicalMemory', '/Value'],
                stdout=subprocess.PIPE, universal_newlines=True, shell=True
            )
            lines = result.stdout.splitlines()
            total = 0
            free = 0
            for line in lines:
                if 'TotalVisibleMemorySize' in line:
                    total = int(line.split('=')[1])
                elif 'FreePhysicalMemory' in line:
                    free = int(line.split('=')[1])
            
            mem['MemTotal'] = total * 1024
            mem['MemFree'] = free * 1024
            
            result = subprocess.run(
                ['wmic', 'pagefile', 'list', 'full'],
                stdout=subprocess.PIPE, universal_newlines=True, shell=True
            )
            lines = result.stdout.splitlines()
            swap_total = 0
            swap_used = 0
            for line in lines:
                if 'AllocatedBaseSize' in line:
                    swap_total = int(line.split('=')[1]) * 1024 * 1024
                elif 'CurrentUsage' in line:
                    swap_used = int(line.split('=')[1]) * 1024 * 1024
            
            mem['SwapTotal'] = swap_total
            mem['SwapFree'] = max(0, swap_total - swap_used)
        else:
            with open('/proc/meminfo', 'r') as f:
                for line in f:
                    parts = line.split()
                    key = parts[0].rstrip(':')
                    if key in ('MemTotal', 'MemFree', 'Buffers', 'Cached', 'SwapTotal', 'SwapFree'):
                        mem[key] = int(parts[1]) * 1024
    except Exception:
        pass
    return mem

def get_root_disk_usage():
    try:
        usage = shutil.disk_usage('/' if not IS_WINDOWS else 'C:\\')
        total_gb = round(usage.total / (1024**3), 2)
        used_gb = round(usage.used / (1024**3), 2)
        free_gb = round(usage.free / (1024**3), 2)
        used_pct = round(calculate_percentage(used_gb, total_gb), 2)
        return {
            'total_gb': total_gb,
            'used_gb': used_gb,
            'free_gb': free_gb,
            'used_percentage': used_pct
        }
    except Exception:
        return {
            'total_gb': None, 'used_gb': None, 'free_gb': None, 'used_percentage': None
        }

def get_top_memory_processes(limit=20):
    try:
        if IS_WINDOWS:
            result = subprocess.run(
                ['wmic', 'process', 'get', 'ProcessId,Name,WorkingSetSize,PercentProcessorTime', '/format:csv'],
                stdout=subprocess.PIPE, universal_newlines=True, shell=True
            )
            lines = result.stdout.strip().splitlines()[1:]
            procs = []
            for line in lines:
                parts = line.split(',')
                if len(parts) >= 4:
                    pid, name, mem, cpu = parts[1], parts[2], parts[3], parts[4] if len(parts) > 4 else "0"
                    try:
                        mem_mb = int(mem) / (1024 * 1024)
                        cpu_pct = float(cpu) if cpu else 0.0
                        procs.append({
                            'pid': pid,
                            'command': name,
                            'mem_percent': mem_mb,
                            'cpu_percent': cpu_pct
                        })
                    except ValueError:
                        continue
            procs.sort(key=lambda x: x['mem_percent'], reverse=True)
            return procs[:limit]
        else:
            result = subprocess.run(
                ['ps', 'axo', 'pid,comm,%mem,%cpu', '--sort=-%mem'],
                stdout=subprocess.PIPE, universal_newlines=True
            )
            lines = result.stdout.strip().splitlines()[1:limit+1]
            procs = []
            for line in lines:
                parts = line.split(None, 3)
                if len(parts) == 4:
                    pid, comm, mem, cpu = parts
                    procs.append({
                        'pid': pid,
                        'command': comm,
                        'mem_percent': float(mem),
                        'cpu_percent': float(cpu)
                    })
            return procs
    except Exception:
        return []

def find_largest_files(paths, top_n=30):
    files = []
    for path in paths:
        if not os.path.exists(path):
            continue
        for root, _, filenames in os.walk(path):
            for f in filenames:
                try:
                    full_path = os.path.join(root, f)
                    size = os.path.getsize(full_path)
                    files.append((full_path, size))
                except (OSError, PermissionError):
                    continue
    files.sort(key=lambda x: x[1], reverse=True)
    return files[:top_n]

def get_cpu_usage():
    if IS_WINDOWS:
        try:
            result = subprocess.run(
                ["wmic", "cpu", "get", "LoadPercentage", "/VALUE"],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                universal_newlines=True, shell=True
            )
            out = result.stdout or result.stderr or ""
            m = re.search(r"LoadPercentage\s*=\s*(\d+)", out, re.IGNORECASE)
            if m:
                return float(m.group(1))

            result = subprocess.run(
                ["wmic", "cpu", "get", "LoadPercentage"],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                universal_newlines=True, shell=True
            )
            for line in (result.stdout or "").splitlines():
                if line.strip().isdigit():
                    return float(line.strip())
        except Exception:
            pass
        return 0.0

    if IS_JETSON:
        try:
            result = subprocess.run(
                ["tegrastats", "--interval", "200", "--count", "1"],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                universal_newlines=True
            )
            out = result.stdout
            m = re.search(r"CPU\s+\[([^\]]+)\]", out)
            if m:
                values = []
                for part in m.group(1).split(","):
                    pm = re.search(r"(\d+)%", part)
                    if pm:
                        values.append(int(pm.group(1)))
                if values:
                    return round(sum(values) / len(values), 2)
        except Exception:
            pass
        if psutil:
            return round(psutil.cpu_percent(interval=0.5), 2)
        return 0.0

    if IS_LINUX:
        if psutil:
            try:
                usage = psutil.cpu_percent(interval=0.5)
                return round(usage, 2)
            except Exception:
                pass

        try:
            def read_proc_stat():
                with open("/proc/stat", "r") as f:
                    for line in f:
                        if line.startswith("cpu "):
                            nums = [int(x) for x in line.split()[1:]]
                            idle = nums[3] + nums[4]
                            total = sum(nums)
                            return idle, total
                return None, None

            idle1, total1 = read_proc_stat()
            time.sleep(0.3)
            idle2, total2 = read_proc_stat()
            if idle1 and total1 and (total2 - total1) > 0:
                busy = (total2 - total1) - (idle2 - idle1)
                return round((busy / (total2 - total1)) * 100.0, 2)
        except Exception:
            pass
        return 0.0

    return 0.0

def get_ram_usage():
    try:
        meminfo = read_meminfo()
        if not meminfo:
            return 0
            
        if IS_WINDOWS:
            total = meminfo.get('MemTotal', 0)
            free = meminfo.get('MemFree', 0)
            if total > 0:
                return round((total - free) / total * 100, 2)
            return 0
        else:
            total = meminfo.get('MemTotal', 0)
            free = meminfo.get('MemFree', 0)
            buffers = meminfo.get('Buffers', 0)
            cached = meminfo.get('Cached', 0)
            if total > 0:
                used = total - free - buffers - cached
                return round(used / total * 100, 2)
            return 0
    except Exception:
        return 0

def get_gpu_usage():
    try:
        result = subprocess.run(
            ["nvidia-smi", "--query-gpu=utilization.gpu", "--format=csv,noheader,nounits"],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True
            )
        if result.returncode == 0:
            vals = [int(v.strip()) for v in result.stdout.splitlines() if v.strip().isdigit()]
            if vals:
                return round(sum(vals) / len(vals), 2)
    except FileNotFoundError:
        pass

    if IS_JETSON:
        try:
            result = subprocess.run(
                ["tegrastats", "--interval", "200", "--count", "1"],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True
            )
            out = result.stdout
            m = re.search(r"GR3D_FREQ\s+(\d+)%", out)
            if m:
                return float(m.group(1))
        except Exception:
            pass

        try:
            import jtop
            with jtop.jtop() as jetson:
                if jetson.ok():
                    gpu_load = jetson.stats.get("GPU", {}).get("val", None)
                    if gpu_load is not None:
                        return float(gpu_load)
        except Exception:
            pass
        return 0.0

    return 0.0

def get_temperature():
    try:
        if IS_WINDOWS:
            result = subprocess.run(
                ['wmic', '/namespace:\\\\root\\wmi', 'path', 'MSAcpi_ThermalZoneTemperature', 'get', 'CurrentTemperature'],
                stdout=subprocess.PIPE, universal_newlines=True, shell=True
            )
            lines = result.stdout.splitlines()
            for line in lines:
                if line.strip().isdigit():
                    return (int(line.strip()) / 10) - 273.15
            return 0
        else:
            temp_paths = [
                '/sys/class/thermal/thermal_zone0/temp',
                '/sys/class/hwmon/hwmon0/temp1_input',
                '/sys/class/hwmon/hwmon1/temp1_input'
            ]
            for path in temp_paths:
                if os.path.exists(path):
                    with open(path, 'r') as f:
                        temp = float(f.read().strip())
                        return temp / 1000.0
            return 0
    except Exception:
        return 0

def collect_system_info():
    system_info = {
        'cpu': get_cpu_usage(),
        'ram': get_ram_usage(),
        'gpu': get_gpu_usage(),
        'temp': get_temperature(),
        'ip_address': get_ip_address(),
        'systemd_services': get_systemd_services()
    }
    
    system_info['platform'] = {
        'system': platform.system(),
        'release': platform.release(),
        'version': platform.version(),
        'machine': platform.machine(),
        'processor': platform.processor()
    }
    
    return system_info

def compute_performance_status(system_info):
    alerts = []
    thresholds = {
        'cpu': 100,
        'ram': 85,
        'disk': 90,
        'gpu': 100,
        'temp': 99
    }
    perf_status = {}
    cpu_usage = system_info['cpu']
    if cpu_usage > thresholds['cpu']:
        alerts.append(f"High CPU usage: {cpu_usage}%")
    perf_status['cpu'] = cpu_usage

    ram_usage = system_info['ram']
    if ram_usage > thresholds['ram']:
        alerts.append(f"High RAM usage: {ram_usage}%")
    perf_status['ram'] = ram_usage

    gpu_usage = system_info['gpu']
    if gpu_usage > thresholds['gpu']:
        alerts.append(f"High GPU usage: {gpu_usage}%")
    perf_status['gpu'] = gpu_usage

    temperature = system_info['temp']
    if temperature > thresholds['temp']:
        alerts.append(f"High temperature: {temperature}°C")
    perf_status['temp'] = temperature

    meminfo = read_meminfo()
    swap_total = meminfo.get('SwapTotal', 0)
    if swap_total > 0:
        swap_free = meminfo.get('SwapFree', 0)
        swap_used = swap_total - swap_free
        swap_used_percentage = calculate_percentage(swap_used, swap_total)
        if swap_used_percentage > 85:
            alerts.append(f"High swap usage: {swap_used_percentage:.2f}%")
        perf_status['swap'] = swap_used_percentage

    return perf_status, thresholds, alerts

def get_systemd_service_status(service_names=None):
    """Get detailed status for specific systemd services"""
    if IS_WINDOWS:
        return {}
    
    service_status = {}
    if service_names is None:
        service_names = []
    
    for service in service_names:
        try:
            result = subprocess.run(
                ['systemctl', 'show', service, '--no-pager'],
                stdout=subprocess.PIPE, 
                stderr=subprocess.PIPE,
                universal_newlines=True
            )
            
            status_info = {}
            if result.returncode == 0:
                for line in result.stdout.splitlines():
                    if '=' in line:
                        key, value = line.split('=', 1)
                        status_info[key] = value
                
                service_status[service] = {
                    'active_state': status_info.get('ActiveState', 'unknown'),
                    'sub_state': status_info.get('SubState', 'unknown'),
                    'load_state': status_info.get('LoadState', 'unknown'),
                    'unit_file_state': status_info.get('UnitFileState', 'unknown'),
                    'memory_usage': status_info.get('MemoryCurrent', 'N/A'),
                    'cpu_usage': status_info.get('CPUUsageNSec', 'N/A')
                }
            else:
                service_status[service] = {
                    'active_state': 'not-found',
                    'sub_state': 'not-found',
                    'load_state': 'not-found',
                    'unit_file_state': 'not-found',
                    'memory_usage': 'N/A',
                    'cpu_usage': 'N/A'
                }
                
        except Exception as e:
            print(f"Error checking service {service}: {e}")
            service_status[service] = {
                'active_state': 'error',
                'sub_state': 'error',
                'load_state': 'error',
                'unit_file_state': 'error',
                'memory_usage': 'N/A',
                'cpu_usage': 'N/A'
            }
    
    return service_status

"""def send_to_dynamodb(meta, system_info, perf_status, camera_ip_status, services_proc, python_proc, system_ip_status):
    
    if not dynamodb:
        print("DynamoDB not configured, skipping database update")
        return
    
    timestamp = datetime.utcnow().isoformat() + "Z"
    
    try:
        # Update projects table using DynamoDB JSON format
        projects_table.put_item(
            Item={
                'project_name': PROJECT,
                'AIC_ID': AIC_ID,
                'AIC_NAME': AIC_NAME,
                'customer_name': CUSTOMER_NAME,
                'description': DESCRIPTION,
                'last_updated': timestamp,
                'location': LOCATION,
                'PEER_IPS': IP_LIST,
                'started_on': '',  # You might want to set this to a specific date
                'SYSTEM': {
                    'IP': system_info.get('ip_address', 'N/A'),
                    'NAME': SYSTEM_NAME
                },
                'status': AIC_STATUS,
                'cpu_usage': Decimal(str(perf_status.get('cpu', 0))),
                'ram_usage': Decimal(str(perf_status.get('ram', 0))),
                'gpu_usage': Decimal(str(perf_status.get('gpu', 0))),
                'temperature': Decimal(str(perf_status.get('temp', 0)))
            }
        )
        print(f"Updated projects table for {PROJECT}")
    except Exception as e:
        print(f"Error updating projects table: {e}")
    
    try:
        # Update cameras table using DynamoDB JSON format
        for ip, camera_info in camera_ip_status.items():
            cameras_table.put_item(
                Item={
                    'project_name': PROJECT,
                    'camera_ip': ip,
                    'camera_id': f"CAM_{ip.replace('.', '_')}",
                    'name': camera_info['name'],
                    'status': camera_info['status'],
                    'last_seen': timestamp,
                    'issue': None if camera_info['status'] == 'running' else 'Camera offline'
                }
            )
        print(f"Updated cameras table with {len(camera_ip_status)} cameras")
    except Exception as e:
        print(f"Error updating cameras table: {e}")
    
    try:
        # Update services table using DynamoDB JSON format
        # Include both services and Python files
        all_services = {**services_proc, **python_proc}
        
        for service_name, status_info in all_services.items():
            # Get CPU and memory usage for the service
            cpu_usage = Decimal('0')
            memory_usage = Decimal('0')
            issue = None
            
            if status_info['status'] != 'running':
                issue = 'Service not running'
            else:
                issue = 'Service running'
            services_table.put_item(
                Item={
                    'project_name': PROJECT,
                    'service_name': service_name,
                    'status': status_info['status'],
                    'uptime': status_info.get('uptime', 'N/A'),
                    'cpu': cpu_usage,
                    'memory': memory_usage,
                    'issue': issue,
                    'last_updated': timestamp
                }
            )
        print(f"Updated services table with {len(all_services)} services and Python files")
    except Exception as e:
        print(f"Error updating services table: {e}")
"""

def send_to_dynamodb(meta, system_info, perf_status, camera_ip_status, services_proc, python_proc, system_ip_status):
    """Update data in DynamoDB tables instead of inserting new items"""
    if not dynamodb:
        print("DynamoDB not configured, skipping database update")
        return
    
    timestamp = datetime.utcnow().isoformat() + "Z"
    
    try:
        # First try to get the item to see what key structure it has
        try:
            response = projects_table.get_item(
                Key={
                    'project_name': PROJECT
                }
            )
            item = response.get('Item', {})
            
            if item:
                # Item exists, update it
                projects_table.update_item(
                    Key={
                        'project_name': PROJECT
                    },
                    UpdateExpression="SET #name = :name, customer_name = :customer_name, description = :description, "
                                   "last_updated = :last_updated, #loc = :location, PEER_IPS = :peer_ips, "
                                   "#sys = :system, #status = :status, cpu_usage = :cpu_usage, "
                                   "ram_usage = :ram_usage, gpu_usage = :gpu_usage, temperature = :temperature",
                    ExpressionAttributeNames={
                        '#name': 'AIC_NAME',
                        '#status': 'status',
                        '#loc': 'location',
                        '#sys': 'SYSTEM'
                    },
                    ExpressionAttributeValues={
                        ':name': AIC_NAME,
                        ':customer_name': CUSTOMER_NAME,
                        ':description': DESCRIPTION,
                        ':last_updated': timestamp,
                        ':location': LOCATION,
                        ':peer_ips': IP_LIST,
                        ':system': {
                            'IP': system_info.get('ip_address', 'N/A'),
                            'NAME': SYSTEM_NAME
                        },
                        ':status': AIC_STATUS,
                        ':cpu_usage': Decimal(str(perf_status.get('cpu', 0))),
                        ':ram_usage': Decimal(str(perf_status.get('ram', 0))),
                        ':gpu_usage': Decimal(str(perf_status.get('gpu', 0))),
                        ':temperature': Decimal(str(perf_status.get('temp', 0)))
                    },
                    ReturnValues="UPDATED_NEW"
                )
                print(f"Updated projects table for {PROJECT}")
            else:
                # Item doesn't exist, create it
                projects_table.put_item(
                    Item={
                        'project_name': PROJECT,
                        'AIC_ID': AIC_ID,
                        'AIC_NAME': AIC_NAME,
                        'customer_name': CUSTOMER_NAME,
                        'description': DESCRIPTION,
                        'last_updated': timestamp,
                        'location': LOCATION,
                        'PEER_IPS': IP_LIST,
                        'SYSTEM': {
                            'IP': system_info.get('ip_address', 'N/A'),
                            'NAME': SYSTEM_NAME
                        },
                        'status': AIC_STATUS,
                        'cpu_usage': Decimal(str(perf_status.get('cpu', 0))),
                        'ram_usage': Decimal(str(perf_status.get('ram', 0))),
                        'gpu_usage': Decimal(str(perf_status.get('gpu', 0))),
                        'temperature': Decimal(str(perf_status.get('temp', 0)))
                    }
                )
                print(f"Created new project entry for {PROJECT}")
                
        except Exception as e:
            print(f"Error checking/updating projects table: {e}")
            # Fallback: try with AIC_ID as key
            try:
                projects_table.update_item(
                    Key={
                        'AIC_ID': AIC_ID
                    },
                    UpdateExpression="SET project_name = :project_name, #name = :name, customer_name = :customer_name, "
                                   "description = :description, last_updated = :last_updated, #loc = :location, "
                                   "PEER_IPS = :peer_ips, #sys = :system, #status = :status, cpu_usage = :cpu_usage, "
                                   "ram_usage = :ram_usage, gpu_usage = :gpu_usage, temperature = :temperature",
                    ExpressionAttributeNames={
                        '#name': 'AIC_NAME',
                        '#status': 'status',
                        '#loc': 'location',
                        '#sys': 'SYSTEM'
                    },
                    ExpressionAttributeValues={
                        ':project_name': PROJECT,
                        ':name': AIC_NAME,
                        ':customer_name': CUSTOMER_NAME,
                        ':description': DESCRIPTION,
                        ':last_updated': timestamp,
                        ':location': LOCATION,
                        ':peer_ips': IP_LIST,
                        ':system': {
                            'IP': system_info.get('ip_address', 'N/A'),
                            'NAME': SYSTEM_NAME
                        },
                        ':status': AIC_STATUS,
                        ':cpu_usage': Decimal(str(perf_status.get('cpu', 0))),
                        ':ram_usage': Decimal(str(perf_status.get('ram', 0))),
                        ':gpu_usage': Decimal(str(perf_status.get('gpu', 0))),
                        ':temperature': Decimal(str(perf_status.get('temp', 0)))
                    },
                    ReturnValues="UPDATED_NEW"
                )
                print(f"Updated projects table using AIC_ID as key for {AIC_ID}")
            except Exception as e2:
                print(f"Error updating projects table with AIC_ID key: {e2}")
    
    except Exception as e:
        print(f"Error updating projects table: {e}")
    
    try:
        # Update cameras table - update existing cameras or create new ones
        for ip, camera_info in camera_ip_status.items():
            try:
                # Try to update existing camera
                cameras_table.update_item(
                    Key={
                        'camera_ip': ip
                    },
                    UpdateExpression="SET project_name = :project_name, #name = :name, #status = :status, last_seen = :last_seen, "
                                   "issue = :issue",
                    ExpressionAttributeNames={
                        '#name': 'name',
                        '#status': 'status'
                    },
                    ExpressionAttributeValues={
                        ':project_name': PROJECT,
                        ':name': camera_info['name'],
                        ':status': camera_info['status'],
                        ':last_seen': timestamp,
                        ':issue': None if camera_info['status'] == 'running' else 'Camera offline'
                    },
                    ReturnValues="UPDATED_NEW"
                )
            except ClientError as e:
                if e.response['Error']['Code'] == 'ValidationException':
                    # Camera doesn't exist, create a new one
                    cameras_table.put_item(
                        Item={
                            'project_name': PROJECT,
                            'camera_ip': ip,
                            'camera_id': f"CAM_{ip.replace('.', '_')}",
                            'name': camera_info['name'],
                            'status': camera_info['status'],
                            'last_seen': timestamp,
                            'issue': None if camera_info['status'] == 'running' else 'Camera offline'
                        }
                    )
                else:
                    raise e
        print(f"Updated cameras table with {len(camera_ip_status)} cameras")
    except Exception as e:
        print(f"Error updating cameras table: {e}")
    
    try:
        # Update services table - update existing services or create new ones
        # Include both services and Python files
        all_services = {**services_proc, **python_proc}
        
        for service_name, status_info in all_services.items():
            # Get CPU and memory usage for the service
            cpu_usage = Decimal('0')
            memory_usage = Decimal('0')
            issue = None
            
            if status_info['status'] != 'running':
                issue = 'Service not running'
            else:
                issue = 'Service running'
                
            try:
                # Try to update existing service
                services_table.update_item(
                    Key={
                        'service_name': service_name
                    },
                    UpdateExpression="SET project_name = :project_name, #status = :status, uptime = :uptime, cpu = :cpu, "
                                   "memory = :memory, issue = :issue, last_updated = :last_updated",
                    ExpressionAttributeNames={
                        '#status': 'status'
                    },
                    ExpressionAttributeValues={
                        ':project_name': PROJECT,
                        ':status': status_info['status'],
                        ':uptime': status_info.get('uptime', 'N/A'),
                        ':cpu': cpu_usage,
                        ':memory': memory_usage,
                        ':issue': issue,
                        ':last_updated': timestamp
                    },
                    ReturnValues="UPDATED_NEW"
                )
            except ClientError as e:
                if e.response['Error']['Code'] == 'ValidationException':
                    # Service doesn't exist, create a new one
                    services_table.put_item(
                        Item={
                            'project_name': PROJECT,
                            'service_name': service_name,
                            'status': status_info['status'],
                            'uptime': status_info.get('uptime', 'N/A'),
                            'cpu': cpu_usage,
                            'memory': memory_usage,
                            'issue': issue,
                            'last_updated': timestamp
                        }
                    )
                else:
                    raise e
        print(f"Updated services table with {len(all_services)} services and Python files")
    except Exception as e:
        print(f"Error updating services table: {e}")

def dynamodb_worker():
    """Background worker to send data to DynamoDB every 20 seconds"""
    while not stop_event.is_set():
        try:
            # Collect minimal data for DynamoDB
            system_info = collect_system_info()
            camera_ip_status = check_ip_status(CAMERA_IPS)
            services_proc = check_services_by_name(SERVICES)
            python_proc = check_services_by_name(PYTHON_FILES)
            system_ip_status = check_ip_status({SYSTEM_IP: SYSTEM_NAME})[SYSTEM_IP]
            perf_status, _, _ = compute_performance_status(system_info)
            
            # Prepare metadata
            meta = {
                'aic_id': AIC_ID,
                'aic_name': AIC_NAME,
                'aic_status': AIC_STATUS,
                'timestamp': datetime.now().isoformat(),
                'project': PROJECT
            }
            
            # Send to DynamoDB
            send_to_dynamodb(meta, system_info, perf_status, camera_ip_status, services_proc, python_proc, system_ip_status)
            
            # Wait for 20 seconds or until stopped
            stop_event.wait(20)
        except Exception as e:
            print(f"Error in DynamoDB worker: {e}")
            # Wait a bit before retrying even if there was an error
            stop_event.wait(20)

def format_excel_worksheet(ws):
    # Apply formatting to all cells
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:  # Only format cells with values
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = LEFT_ALIGN
    
    # Auto-adjust column widths (skip merged cells)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not cell.coordinate in ws.merged_cells:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 to avoid overly wide columns
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width

def add_section_header(ws, title, row_num):
    """Add a formatted section header to the worksheet"""
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
    cell = ws.cell(row=row_num, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = CENTER_ALIGN
    return row_num + 1

def add_table_header(ws, headers, row_num):
    """Add a formatted table header to the worksheet"""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
    return row_num + 1

def highlight_warnings(ws, start_row, end_row, col_num, threshold, comparison=">"):
    """Highlight cells that exceed threshold values"""
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col_num)
        try:
            value = float(cell.value)
            if (comparison == ">" and value > threshold) or (comparison == "<" and value < threshold):
                cell.fill = WARNING_FILL
                cell.font = HIGHLIGHT_FONT
        except (ValueError, TypeError):
            pass

def make_excel_report(path, meta, system_info, perf_status, thresholds, alerts,
                      peer_ip_status, camera_ip_status, services_proc, python_proc,
                      system_ip_status, swap_info, root_disk, top_procs, largest_files):
    wb = Workbook()
    ws = wb.active
    ws.title = "Health Report"
    
    current_row = 1
    
    # System Information Header
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    header_cell = ws.cell(row=current_row, column=1, value=f"{AIC_NAME} - System Health Report")
    header_cell.font = Font(bold=True, size=14, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_cell.alignment = CENTER_ALIGN
    current_row += 2
    
    # System Metadata
    current_row = add_section_header(ws, "SYSTEM INFORMATION", current_row)
    ws.cell(row=current_row, column=1, value="AIC ID").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=meta['aic_id'])
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="AIC Name").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=meta['aic_name'])
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="Timestamp").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=meta['timestamp'])
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="Project").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=meta['project'])
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="IP Address").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=system_info.get('ip_address', 'N/A'))
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="Platform").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f"{system_info['platform']['system']} {system_info['platform']['release']}")
    current_row += 2
    
    # Performance Metrics
    current_row = add_section_header(ws, "PERFORMANCE METRICS", current_row)
    current_row = add_table_header(ws, ["Metric", "Value", "Threshold", "Status"], current_row)
    
    metrics = [
        ("CPU Usage", perf_status.get('cpu', 0), thresholds['cpu'], "OK" if perf_status.get('cpu', 0) <= thresholds['cpu'] else "WARNING"),
        ("RAM Usage", perf_status.get('ram', 0), thresholds['ram'], "OK" if perf_status.get('ram', 0) <= thresholds['ram'] else "WARNING"),
        ("GPU Usage", perf_status.get('gpu', 0), thresholds['gpu'], "OK" if perf_status.get('gpu', 0) <= thresholds['gpu'] else "WARNING"),
        ("Temperature", perf_status.get('temp', 0), thresholds['temp'], "OK" if perf_status.get('temp', 0) <= thresholds['temp'] else "WARNING"),
        ("Disk Usage", root_disk.get('used_percentage', 0), thresholds['disk'], "OK" if root_disk.get('used_percentage', 0) <= thresholds['disk'] else "WARNING")
    ]
    
    # Add swap usage if available
    if 'swap' in perf_status:
        metrics.append(("Swap Usage", perf_status.get('swap', 0), 85, "OK" if perf_status.get('swap', 0) <= 85 else "WARNING"))
    
    for metric, value, threshold, status in metrics:
        ws.cell(row=current_row, column=1, value=metric)
        ws.cell(row=current_row, column=2, value=f"{value}%")
        ws.cell(row=current_row, column=3, value=f"{threshold}%")
        status_cell = ws.cell(row=current_row, column=4, value=status)
        
        if status == "WARNING":
            status_cell.fill = WARNING_FILL
            status_cell.font = HIGHLIGHT_FONT
        else:
            status_cell.fill = SUCCESS_FILL
            status_cell.font = SUCCESS_FONT
            
        current_row += 1
    
    current_row += 1
    
    # Alerts Section
    if alerts:
        current_row = add_section_header(ws, "ALERTS", current_row)
        for alert in alerts:
            alert_cell = ws.cell(row=current_row, column=1, value=alert)
            alert_cell.fill = WARNING_FILL
            alert_cell.font = HIGHLIGHT_FONT
            current_row += 1
        current_row += 1
    
    # Network Status
    current_row = add_section_header(ws, "NETWORK STATUS", current_row)
    
    # System IP
    ws.cell(row=current_row, column=1, value="System IP").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=SYSTEM_IP)
    status_cell = ws.cell(row=current_row, column=3, value=system_ip_status.get('status', 'unknown'))
    if system_ip_status.get('status') == 'running':
        status_cell.fill = ONLINE_FILL
        status_cell.font = SUCCESS_FONT
    else:
        status_cell.fill = WARNING_FILL
        status_cell.font = HIGHLIGHT_FONT
    current_row += 1
    
    current_row += 1
    
    # Camera IPs
    current_row = add_table_header(ws, ["Camera Name", "IP Address", "Status"], current_row)
    for ip, val in camera_ip_status.items():
        ws.cell(row=current_row, column=1, value=val['name'])
        ws.cell(row=current_row, column=2, value=ip)
        status_cell = ws.cell(row=current_row, column=3, value=val['status'])
        if val['status'] == 'running':
            status_cell.fill = ONLINE_FILL
            status_cell.font = SUCCESS_FONT
        else:
            status_cell.fill = WARNING_FILL
            status_cell.font = HIGHLIGHT_FONT
        current_row += 1
    
    current_row += 2
    
    # Services Status
    current_row = add_section_header(ws, "SERVICES STATUS", current_row)
    current_row = add_table_header(ws, ["Service Name", "Status", "Uptime"], current_row)
    
    for name, status_info in services_proc.items():
        ws.cell(row=current_row, column=1, value=name)
        status_cell = ws.cell(row=current_row, column=2, value=status_info['status'])
        ws.cell(row=current_row, column=3, value=status_info.get('uptime', 'N/A'))
        
        if status_info['status'] == 'running':
            status_cell.fill = SUCCESS_FILL
            status_cell.font = SUCCESS_FONT
        else:
            status_cell.fill = WARNING_FILL
            status_cell.font = HIGHLIGHT_FONT
        current_row += 1
    
    current_row += 1
    
    # Python Processes
    current_row = add_table_header(ws, ["Python File", "Status", "Uptime"], current_row)
    
    for name, status_info in python_proc.items():
        ws.cell(row=current_row, column=1, value=name)
        status_cell = ws.cell(row=current_row, column=2, value=status_info['status'])
        ws.cell(row=current_row, column=3, value=status_info.get('uptime', 'N/A'))
        
        if status_info['status'] == 'running':
            status_cell.fill = SUCCESS_FILL
            status_cell.font = SUCCESS_FONT
        else:
            status_cell.fill = WARNING_FILL
            status_cell.font = HIGHLIGHT_FONT
        current_row += 1
    
    current_row += 2
    
    # Resource Usage
    current_row = add_section_header(ws, "RESOURCE USAGE", current_row)
    
    # Swap Info
    ws.cell(row=current_row, column=1, value="Swap Total").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f"{swap_info.get('swap_total_mb', 0)} MB")
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="Swap Used").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f"{swap_info.get('swap_used_mb', 0)} MB")
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="Swap Usage").font = Font(bold=True)
    usage_cell = ws.cell(row=current_row, column=2, value=f"{swap_info.get('swap_used_percentage', 0)}%")
    if swap_info.get('swap_used_percentage', 0) > 50:
        usage_cell.fill = WARNING_FILL
        usage_cell.font = HIGHLIGHT_FONT
    else:
        usage_cell.fill = SUCCESS_FILL
        usage_cell.font = SUCCESS_FONT
    current_row += 1
    
    # Disk Usage
    ws.cell(row=current_row, column=1, value="Disk Total").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f"{root_disk.get('total_gb', 0)} GB")
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="Disk Used").font = Font(bold=True)
    ws.cell(row=current_row, column=2, value=f"{root_disk.get('used_gb', 0)} GB")
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="Disk Usage").font = Font(bold=True)
    usage_cell = ws.cell(row=current_row, column=2, value=f"{root_disk.get('used_percentage', 0)}%")
    if root_disk.get('used_percentage', 0) > thresholds['disk']:
        usage_cell.fill = WARNING_FILL
        usage_cell.font = HIGHLIGHT_FONT
    else:
        usage_cell.fill = SUCCESS_FILL
        usage_cell.font = SUCCESS_FONT
    current_row += 2
    
    current_row = add_section_header(ws, "TOP PROCESSES BY MEMORY", current_row)
    current_row = add_table_header(ws, ["PID", "Command", "Memory %", "CPU %"], current_row)
    
    for proc in top_procs:
        ws.cell(row=current_row, column=1, value=proc['pid'])
        ws.cell(row=current_row, column=2, value=proc['command'])
        ws.cell(row=current_row, column=3, value=f"{proc['mem_percent']:.1f}%")
        ws.cell(row=current_row, column=4, value=f"{proc['cpu_percent']:.1f}%")
        current_row += 1
    
    current_row += 2
    
    # Largest Files
    current_row = add_section_header(ws, "LARGEST FILES", current_row)
    current_row = add_table_header(ws, ["File Path", "Size (MB)"], current_row)
    
    for file_path, size_bytes in largest_files:
        size_mb = size_bytes / (1024 * 1024)
        ws.cell(row=current_row, column=1, value=file_path)
        ws.cell(row=current_row, column=2, value=f"{size_mb:.1f}")
        current_row += 1

    # Systemd Services Status
    current_row = add_section_header(ws, "SYSTEMD SERVICES STATUS", current_row)
    current_row = add_table_header(ws, ["Service Name", "Load State", "Active State", "Sub State"], current_row)
    
    for service in system_info.get('systemd_services', []):
        ws.cell(row=current_row, column=1, value=service.get('service_name', 'N/A'))
        ws.cell(row=current_row, column=2, value=service.get('load_state', 'N/A'))
        
        active_state = service.get('active_state', 'N/A')
        active_cell = ws.cell(row=current_row, column=3, value=active_state)
        
        sub_state = service.get('sub_state', 'N/A')
        sub_cell = ws.cell(row=current_row, column=4, value=sub_state)
        
        # Highlight inactive or failed services
        if active_state not in ['active', 'activating']:
            active_cell.font = HIGHLIGHT_FONT
            sub_cell.font = HIGHLIGHT_FONT
        elif sub_state == 'failed':
            active_cell.fill = WARNING_FILL
            active_cell.font = HIGHLIGHT_FONT
            sub_cell.fill = WARNING_FILL
            sub_cell.font = HIGHLIGHT_FONT
        else:
            active_cell.fill = SUCCESS_FILL
            active_cell.font = SUCCESS_FONT
            sub_cell.fill = SUCCESS_FILL
            sub_cell.font = SUCCESS_FONT
            
        current_row += 1
    
    current_row += 2
    
    # Apply final formatting
    format_excel_worksheet(ws)
    
    # Save the workbook
    wb.save(path)
    print(f"Excel report saved to: {path}")

def send_email(subject, body, attachment_path=None):
    """Send email with optional attachment"""
    if not SENDER_EMAIL or not PASSWORD or not RECEIVER_EMAILS:
        print("Email configuration missing. Skipping email notification.")
        return False
    
    try:
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = SENDER_EMAIL
        msg['To'] = ', '.join(RECEIVER_EMAILS)
        msg.set_content(body)
        
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, 'rb') as f:
                file_data = f.read()
                file_name = os.path.basename(attachment_path)
                msg.add_attachment(file_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=file_name)
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(SENDER_EMAIL, PASSWORD)
            smtp.send_message(msg)
        
        print("Email sent successfully!")
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

def should_send_alert(perf_status, thresholds, services_proc, python_proc, systemd_service_status):
    """
    Check if we need to send an alert based on system conditions
    Returns True if any issue is detected, False otherwise
    """
    # Check performance metrics
    if (perf_status.get('cpu', 0) > thresholds['cpu'] or
        perf_status.get('ram', 0) > thresholds['ram'] or
        perf_status.get('gpu', 0) > thresholds['gpu'] or
        perf_status.get('temp', 0) > thresholds['temp']):
        return True
    
    # Check disk usage
    root_disk = get_root_disk_usage()
    if root_disk.get('used_percentage', 0) > thresholds['disk']:
        return True
    
    # Check swap usage
    meminfo = read_meminfo()
    swap_total = meminfo.get('SwapTotal', 0)
    if swap_total > 0:  # Only check if swap exists
        swap_free = meminfo.get('SwapFree', 0)
        swap_used = swap_total - swap_free
        swap_used_percentage = calculate_percentage(swap_used, swap_total)
        if swap_used_percentage > 85:  # 85% threshold for swap
            return True
    
    # Check services
    for name, status_info in services_proc.items():
        if status_info['status'] != 'running':
            return True
    
    # Check Python files
    for name, status_info in python_proc.items():
        if status_info['status'] != 'running':
            return True
    
    # Check systemd services (Linux only)
    #if not IS_WINDOWS:
        #for status in systemd_service_status.values():
            #if status.get('active_state') != 'active':
                #return True
    
    # If we reached here, everything is OK
    
    return False

def main():
    global AIC_STATUS, stop_event
    
    print(f"Starting system health check for {AIC_NAME} ({AIC_ID})...")
    
    # Start DynamoDB worker thread
    db_thread = Thread(target=dynamodb_worker, daemon=True)
    db_thread.start()
    print("DynamoDB worker thread started")
    
    try:
        # Collect system information
        system_info = collect_system_info()
        
        # Check IP statuses
        print("Checking network connectivity...")
        peer_ip_status = check_ip_status(IP_LIST)
        camera_ip_status = check_ip_status(CAMERA_IPS)
        system_ip_status = check_ip_status({SYSTEM_IP: SYSTEM_NAME})[SYSTEM_IP]
        
        # Check services
        print("Checking services...")
        services_proc = check_services_by_name(SERVICES)
        python_proc = check_services_by_name(PYTHON_FILES)
        
        # Get detailed systemd service status for important services
        print("Checking systemd services...")
        systemd_service_status = get_systemd_service_status(SERVICES)
        
        # Determine AIC status based on critical services
        critical_services_ok = True
        for service, status_info in services_proc.items():
            if status_info['status'] != 'running':
                critical_services_ok = False
                break
                
        if not IS_WINDOWS:
            for service, status in systemd_service_status.items():
                if status.get('active_state') != 'active':
                    critical_services_ok = False
                    break
        
        AIC_STATUS = "on" if critical_services_ok else "off"
        
        # Get memory and disk info
        meminfo = read_meminfo()
        swap_total = meminfo.get('SwapTotal', 0)
        swap_free = meminfo.get('SwapFree', 0)
        swap_used = swap_total - swap_free
        swap_used_percentage = calculate_percentage(swap_used, swap_total) if swap_total > 0 else 0
        
        swap_info = {
            'swap_total_mb': round(swap_total / (1024 * 1024), 2),
            'swap_used_mb': round(swap_used / (1024 * 1024), 2),
            'swap_used_percentage': round(swap_used_percentage, 2)
        }
        
        root_disk = get_root_disk_usage()
        
        # Get top processes and largest files
        print("Analyzing system resources...")
        top_procs = get_top_memory_processes()
        largest_files = find_largest_files(SEARCH_PATHS)
        
        # Compute performance metrics and alerts
        perf_status, thresholds, alerts = compute_performance_status(system_info)
        
        # Prepare metadata
        meta = {
            'aic_id': AIC_ID,
            'aic_name': AIC_NAME,
            'aic_status': AIC_STATUS,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'project': PROJECT
        }
        
        # Generate Excel report
        report_filename = f"{AIC_NAME.replace(' ', '_')}_Health_Report.xlsx"
        make_excel_report(
            report_filename, meta, system_info, perf_status, thresholds, alerts,
            peer_ip_status, camera_ip_status, services_proc, python_proc,
            system_ip_status, swap_info, root_disk, top_procs, largest_files
        )
        
        # Check if we need to send an alert
        need_alert = should_send_alert(perf_status, thresholds, services_proc, python_proc, systemd_service_status)
        
        if need_alert:
            # Prepare email content
            subject = f"ALERT: {AIC_NAME} Health Issues Detected"
            
            # Count systemd services status
            active_systemd_services = 0
            total_systemd_services = len(systemd_service_status)
            for status in systemd_service_status.values():
                if status.get('active_state') == 'active':
                    active_systemd_services += 1
            
            # Count running services and Python files
            running_services = sum(1 for status_info in services_proc.values() if status_info['status'] == 'running')
            running_python_files = sum(1 for status_info in python_proc.values() if status_info['status'] == 'running')
            
            body = f"""
URGENT: AIC System Health Alert
===============================

System Information:
-------------------
AIC ID: {AIC_ID}
AIC Name: {AIC_NAME}
Timestamp: {meta['timestamp']}
Project: {PROJECT}
IP Address: {system_info.get('ip_address', 'N/A')}
Platform: {system_info['platform']['system']} {system_info['platform']['release']}

Performance Issues:
-------------------
CPU Usage: {perf_status.get('cpu', 0)}% (Threshold: {thresholds['cpu']}%)
RAM Usage: {perf_status.get('ram', 0)}% (Threshold: {thresholds['ram']}%)
GPU Usage: {perf_status.get('gpu', 0)}% (Threshold: {thresholds['gpu']}%)
Temperature: {perf_status.get('temp', 0)}°C (Threshold: {thresholds['temp']}°C)
Disk Usage: {root_disk.get('used_percentage', 0)}% (Threshold: {thresholds['disk']}%)
Swap Usage: {swap_info.get('swap_used_percentage', 0)}% (Threshold: 85%)

Service Status:
---------------
Services Running: {running_services}/{len(services_proc)}
Python Files Running: {running_python_files}/{len(python_proc)}
Systemd Services Active: {active_systemd_services}/{total_systemd_services}

Alerts:
-------
{alerts if alerts else 'No critical alerts'}

See attached Excel report for detailed information.

Immediate attention required!
"""
            
            # Send email with report
            print("Sending alert email notification...")
            send_email(subject, body, report_filename)
        else:
            print("System is healthy. No email notification sent.")
        
        print("System health check completed!")
        
        # Keep the main thread running
        while True:
            time.sleep(1)
            
    except KeyboardInterrupt:
        print("Shutting down...")
        stop_event.set()
        db_thread.join(timeout=5)
        print("Shutdown complete")

if __name__ == "__main__":
    main()

